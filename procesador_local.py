# Sección de importaciones (al inicio del archivo)
from datetime import datetime  # <-- Esta es la importación crítica
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from io import BytesIO, StringIO
import openpyxl
import time
import re

# Configuración
SERVICE_ACCOUNT_FILE = 'credentials.json'
SCOPES = ['https://www.googleapis.com/auth/drive']
FOLDER_ID = '1pkTrRlS-WKJyR7_MVpU_emap2MaXgo_F'
CHECK_INTERVAL = 15  # Segundos entre verificaciones

# Inicialización
creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=creds)

def parse_compact_text(text):
    """Parsea el formato de línea única con comillas opcionales"""
    elements = []
    current = []
    in_quotes = False
    
    # Separar considerando comillas
    for part in text.split(' '):
        if not part:
            continue
        if part.startswith('"') and not in_quotes:
            in_quotes = True
            current.append(part[1:])
        elif part.endswith('"') and in_quotes:
            in_quotes = False
            current.append(part[:-1])
            elements.append(' '.join(current))
            current = []
        elif in_quotes:
            current.append(part)
        else:
            elements.append(part)
    
    # Procesar elementos
    columns = []
    current_col = {}
    i = 0
    n = len(elements)
    
    while i < n:
        if elements[i] == 'Columna' and i+1 < n:
            if current_col:
                columns.append(current_col)
            current_col = {
                'name': elements[i+1],
                'unit': '',
                'notes': '',
                'values': []
            }
            i += 2
        elif elements[i] == 'unidad' and i+1 < n and current_col:
            current_col['unit'] = elements[i+1]
            i += 2
        elif (elements[i] in ['apreciacion', 'apreciación']) and i+1 < n and current_col:
            current_col['notes'] = elements[i+1]
            i += 2
        elif elements[i] == 'datos' and current_col:
            i += 1
            while i < n and elements[i] not in ['Columna', 'unidad', 'apreciacion', 'apreciación']:
                current_col['values'].append(elements[i])
                i += 1
        else:
            i += 1
    
    if current_col:
        columns.append(current_col)
    
    return columns

def parse_multiline_format(text):
    """Procesa el formato multi-línea"""
    columns = []
    current_col = {}
    mode = None
    
    for line in text.split('\n'):
        line = line.strip().lower()
        
        if line.startswith('Columna'):
            if current_col:
                columns.append(current_col)
            name = ' '.join(line.split()[1:])
            current_col = {
                'name': name,
                'unit': '',
                'notes': '',
                'values': []
            }
            mode = None
        
        elif line.startswith('unidad'):
            current_col['unit'] = ' '.join(line.split()[1:])
            mode = None
        
        elif line.startswith('apreciacion') or line.startswith('apreciación'):
            current_col['notes'] = ' '.join(line.split()[1:])
            mode = None
        
        elif line.startswith('datos'):
            mode = 'data'
        
        elif mode == 'data':
            values = [v for v in line.split() if v]
            current_col['values'].extend(values)
    
    if current_col:
        columns.append(current_col)
    
    return columns

def parse_compact_text(text):
    """Procesa el formato de línea única"""
    elements = []
    current = []
    in_quotes = False
    
    # Separar considerando comillas
    for part in text.split(' '):
        if not part:
            continue
        if part.startswith('"') and not in_quotes:
            in_quotes = True
            current.append(part[1:])
        elif part.endswith('"') and in_quotes:
            in_quotes = False
            current.append(part[:-1])
            elements.append(' '.join(current))
            current = []
        elif in_quotes:
            current.append(part)
        else:
            elements.append(part)
    
    # Procesar elementos
    columns = []
    current_col = {}
    i = 0
    n = len(elements)
    
    while i < n:
        if elements[i] == 'Columna' and i+1 < n:
            if current_col:
                columns.append(current_col)
            current_col = {
                'name': elements[i+1],
                'unit': '',
                'notes': '',
                'values': []
            }
            i += 2
        elif elements[i] == 'unidad' and i+1 < n and current_col:
            current_col['unit'] = elements[i+1]
            i += 2
        elif (elements[i] in ['apreciacion', 'apreciación']) and i+1 < n and current_col:
            current_col['notes'] = elements[i+1]
            i += 2
        elif elements[i] == 'datos' and current_col:
            i += 1
            while i < n and elements[i] not in ['Columna', 'unidad', 'apreciacion', 'apreciación']:
                current_col['values'].append(elements[i])
                i += 1
        else:
            i += 1
    
    if current_col:
        columns.append(current_col)
    
    return columns
def parse_text_to_columns(text):
    """
    Procesa texto con múltiples columnas identificadas por la palabra 'columna'
    Ejemplo de formato:
    "columna Edad datos 25 30 28 columna Nombre datos Juan Pedro Ana"
    """
    columns = []
    current_col = None
    words = text.split()
    i = 0
    
    while i < len(words):
        if words[i].lower() == 'columna' and i+1 < len(words):
            if current_col:  # Guardar la columna anterior
                columns.append(current_col)
            # Nueva columna
            current_col = {
                'name': words[i+1],
                'unit': '',
                'notes': '',
                'values': []
            }
            i += 2
        elif current_col is not None:
            if words[i].lower() == 'unidad' and i+1 < len(words):
                current_col['unit'] = words[i+1]
                i += 2
            elif words[i].lower() in ['apreciacion', 'apreciación'] and i+1 < len(words):
                current_col['notes'] = words[i+1]
                i += 2
            elif words[i].lower() == 'datos':
                i += 1
                # Capturar todos los valores hasta la próxima columna
                while i < len(words) and words[i].lower() not in ['columna', 'unidad', 'apreciacion', 'apreciación']:
                    current_col['values'].append(words[i])
                    i += 1
            else:
                i += 1
        else:
            i += 1
    
    if current_col:  # Añadir la última columna procesada
        columns.append(current_col)
    
    return columns

def get_unprocessed_files():
    """Obtiene archivos .txt no procesados"""
    query = f"'{FOLDER_ID}' in parents and mimeType='text/plain' and name contains 'datos_tabla_'"
    results = drive_service.files().list(q=query, fields="files(id,name)").execute()
    return results.get('files', [])

def mark_as_processed(file_id):
    """Renombra el archivo para marcarlo como procesado"""
    drive_service.files().update(
        fileId=file_id,
        body={'name': f"PROCESADO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"},  # Cambiado aquí
        fields='name'
    ).execute()

def process_file(file_id):
    try:
        # Descargar archivo
        request = drive_service.files().get_media(fileId=file_id)
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        content = fh.getvalue().decode('utf-8')
        print(f"\nContenido crudo del archivo:\n{content}\n")  # Debug
        
        # Determinar formato y extraer datos
        if '---DATOS---' in content:
            # Formato multi-línea
            data_section = content.split('---DATOS---')[1].strip()
            print(f"Datos extraídos (multilínea):\n{data_section}\n")  # Debug
            columns = parse_multiline_format(data_section)
        else:
            # Formato línea única
            data_line = content.split('\n')[0].strip()
            print(f"Datos extraídos (línea única):\n{data_line}\n")  # Debug
            columns = parse_text_to_columns(content)
    
        if not columns:
            print("⚠️ No se encontraron columnas válidas en el archivo")
            return None
        
        print(f"Columnas procesadas: {columns}\n")  # Debug
        
        # Crear Excel
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Escribir datos
        for col_idx, col in enumerate(columns, 1):
            # Encabezado
            ws.cell(row=1, column=col_idx, value=col['name'])
            
            # Unidad (fila 2)
            if col['unit']:
                ws.cell(row=2, column=col_idx, value=f"Unidad: {col['unit']}")
            
            # Apreciación (fila 3)
            if col['notes']:
                ws.cell(row=3, column=col_idx, value=f"Apreciación: {col['notes']}")
            
            # Valores (desde fila 4)
            for row_idx, value in enumerate(col['values'], 4):
                try:
                    num_value = float(value)
                    ws.cell(row=row_idx, column=col_idx, value=num_value)
                except ValueError:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar y subir
        wb.save(output)
        output.seek(0)
        
        # Subir a Drive
        excel_name = f"tabla_procesada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_metadata = {
            'name': excel_name,
            'parents': [FOLDER_ID],
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        media = MediaIoBaseUpload(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        uploaded_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,name,webViewLink'
        ).execute()
        
        print(f"Archivo Excel creado: {uploaded_file.get('name')}")
        return uploaded_file.get('name')
    
    except Exception as e:
        print(f"Error en process_file: {str(e)}")
        raise

def main_loop():
    """Bucle principal de monitoreo"""
    processed_files = set()
    
    while True:
        try:
            print(f"\n[{datetime.now()}] Buscando archivos nuevos...")  # Cambiado aquí
            files = get_unprocessed_files()
            
            for file in files:
                if file['id'] not in processed_files:
                    print(f"Procesando {file['name']}...")
                    try:
                        result = process_file(file['id'])
                        if result:
                            mark_as_processed(file['id'])
                            processed_files.add(file['id'])
                            print(f"✅ Generado: {result}")
                    except Exception as e:
                        print(f"❌ Error procesando {file['name']}: {str(e)}")
            
            time.sleep(CHECK_INTERVAL)
            
        except KeyboardInterrupt:
            print("\nDeteniendo el monitor...")
            break
        except Exception as e:
            print(f"Error en el bucle principal: {str(e)}")
            time.sleep(60)

if __name__ == '__main__':
    print(f"Iniciando procesador local a las {datetime.now()}...")  # <-- Usa datetime.now()
    main_loop()